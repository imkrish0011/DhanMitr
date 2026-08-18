import json
from datetime import UTC, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook


URL = (
    "https://nsearchives.nseindia.com//web/mediaattachment/"
    "2026-07/Primary_Market_data__June26_20260717144111.xlsx"
)

OUTPUT = Path(
    "rag/data/periodic/nse/"
    "nse_primary_market_june_2026_raw.json"
)

SOURCE_NAME = "National Stock Exchange of India"


def clean_value(value):
    if value is None:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def main():
    response = requests.get(
        URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    response.raise_for_status()

    temp_file = Path(
        "rag/data/periodic/nse/"
        "nse_primary_market_temp.xlsx"
    )

    temp_file.write_bytes(response.content)

    workbook = load_workbook(
        temp_file,
        read_only=True,
        data_only=True,
    )

    sheet = workbook[workbook.sheetnames[0]]

    rows = []

    for row in sheet.iter_rows():
        values = [
            clean_value(cell.value)
            for cell in row[:14]
        ]

        while values and values[-1] == "":
            values.pop()

        if any(values):
            rows.append(values)

    workbook.close()
    temp_file.unlink(missing_ok=True)

    if len(rows) < 2:
        raise RuntimeError(
            "NSE primary market data is empty."
        )

    headers = rows[0]

    records = []

    for row in rows[1:]:
        record = {}

        for index, header in enumerate(headers):
            if not header:
                continue

            value = row[index] if index < len(row) else ""
            record[str(header)] = value

        records.append(record)

    content_lines = []

    for record in records:
        content_lines.append(
            " | ".join(
                f"{key}: {value}"
                for key, value in record.items()
                if value != ""
            )
        )

    content = "\n".join(content_lines)

    output = {
        "scheme_name": (
            "NSE Primary Market Monthly Report - June 2026"
        ),
        "source_url": (
            "https://www.nseindia.com/static/regulations/"
            "segment-wise-historical-reports-capital-primary-market"
        ),
        "source_name": SOURCE_NAME,
        "retrieved_at": datetime.now(UTC).isoformat(),
        "data_type": "periodic",
        "sections": [
            {
                "heading": (
                    "Primary Market Data - June 2026"
                ),
                "content": content,
            }
        ],
    }

    OUTPUT.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    OUTPUT.write_text(
        json.dumps(
            output,
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )

    print(
        f"Saved raw NSE data to {OUTPUT}"
    )
    print(f"Records: {len(records)}")
    print(f"Content length: {len(content)}")


if __name__ == "__main__":
    main()