import json
from datetime import UTC, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook


URL = (
    "https://www.sebi.gov.in/sebi_data/commondocs/"
    "jul-2026/Monthly%20SEBI%20Bulletin%2027.07.2026_p.xlsx"
)

OUTPUT = Path(
    "rag/data/periodic/sebi/raw/"
    "sebi_bulletin_july_2026_raw.json"
)

SOURCE_NAME = "Securities and Exchange Board of India"

SELECTED_SHEETS = {
    "6": "Table 6: Consolidated Resource Mobilisation by Corporates/Other Entities through Primary Markets",
    "8": "Table 8: Offer for Sale through Exchanges",
    "11": "Table 11: Size-wise Classification of Capital Raised through Public and Rights Issues (Equity)",
    "14": "Table 14: Private Placement of Corporate Debt listed at BSE and NSE",
    "15": "Table 15: Trends in Cash Segment of BSE",
    "16": "Table 16: Trends in Cash Segment of NSE",
    "60": "Table 60: Trends in Resource Mobilization by Mutual Funds",
    "61": "Table 61: Scheme-wise Statistics of Mutual Funds",
    "62": "Table 62: Assets Managed by Portfolio Managers",
}


def clean_value(value):
    if value is None:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def main():
    response = requests.get(
        URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    response.raise_for_status()

    temp_file = Path(
        "rag/data/periodic/sebi/"
        "sebi_bulletin_july_2026_temp.xlsx"
    )

    temp_file.write_bytes(response.content)

    workbook = load_workbook(
        temp_file,
        read_only=True,
        data_only=True,
    )

    sections = []

    for sheet_name, table_name in SELECTED_SHEETS.items():

        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(
                f"Sheet {sheet_name} not found."
            )

        sheet = workbook[sheet_name]

        rows = []

        for row in sheet.iter_rows():
            values = [
                clean_value(cell.value)
                for cell in row[:26]
            ]

            while values and values[-1] == "":
                values.pop()

            if any(values):
                rows.append(values)

        content_lines = []

        for row in rows:
            content_lines.append(" | ".join(row))

        content = "\n".join(content_lines)

        if not content:
            raise RuntimeError(
                f"{table_name} is empty."
            )

        sections.append(
            {
                "heading": table_name,
                "content": content,
            }
        )

    workbook.close()

    temp_file.unlink(missing_ok=True)

    record = {
        "scheme_name": "SEBI Bulletin - July 2026",
        "source_url": (
            "https://www.sebi.gov.in/"
            "reports-and-statistics/publications/"
            "jul-2026/"
            "sebi-bulletin-july-2026_103048.html"
        ),
        "source_name": SOURCE_NAME,
        "retrieved_at": datetime.now(UTC).isoformat(),
        "data_type": "periodic",
        "sections": sections,
    }

    OUTPUT.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    OUTPUT.write_text(
        json.dumps(
            record,
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )

    print(
        f"Saved raw SEBI bulletin data to {OUTPUT}"
    )
    print(f"Sections: {len(sections)}")
    print(
        "Empty sections:",
        sum(
            1
            for section in sections
            if not section["content"].strip()
        ),
    )


if __name__ == "__main__":
    main()